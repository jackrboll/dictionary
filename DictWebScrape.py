#!/usr/bin/env python
# -*- coding: utf-8 -*-

#User inputs word, finnish/english, spreadsheet file

import os
import bs4
import openpyxl
import sys, requests
import logging

logging.basicConfig(filename='dict.log',level=logging.INFO)


COLUMN_NAMES = ['Word', 'Definition', 'Tags', 'Date', 'Count']

###################
## Set up the spreadsheets
##
##
###################
try: 
	spreadsheet = sys.argv[3]
except:
	spreadsheet = ""

def populate_columns():
	if (spreadsheet == ""):
		spreadsheet = 'finnish.xlsx'
		log.info("Creating new spreadsheet 'finnish.xlsx'.")
		wb = openpyxl.Workbook()
		ws = wb.active
		#add column names
		for i in range(0,5):
			ws['A1:E1'][0][i].value = COLUMN_NAMES[i]
	else:
		try:
			wb = openpyxl.load_workbook(spreadsheet)
			ws = wb.active
		except: 
			spreadsheet = 'finnish.xlsx'
			print "No spreadsheet found.  \nCreating new spreadsheet called 'finnish.xlsx'."  
			wb = openpyxl.Workbook()
			ws = wb.active
			#add column names
			for i in range(0,5):
				 ws['A1:E1'][0][i].value = COLUMN_NAMES[i]

##################
## Get the data from the internet
##
##
##################

try: 
	word = sys.argv[1]
except:
	word = 'mulkaista'

#if word contains a unlaut or o umlaut this won't work need to url encode the word first
a_umlaut = "%C3%A4"
o_umlaut = "%C3%B6"
a_swedish = "%C3%A5"

word.replace("ä", a_umlaut)
word.replace("ö", o_umlaut)
word.replace("å", a_swedish)

try:
	lang = sys.argv[2] 
except:
	lang = 'finnish'

if lang == 'finnish':
	url = 'https://redfoxsanakirja.fi/sanakirja/-/s/fin/eng/' + word
else:
	#assuming lang is english
	url = 'https://redfoxsanakirja.fi/sanakirja/-/s/eng/fin/' + word

res = requests.get(url)
soup = bs4.BeautifulSoup(res.text)
soup.find_all(class_='translation-table')[0].find_all(class_='single-word')[0].strings



#get definitions into usable list 
definitions = []
for i in range(0, len(soup.find_all(class_='translation-table')[0].select('li'))):
	definition =""
	for string in (soup.find_all(class_='translation-table')[0].select('li')[i].strings):
		definition += string
		definition.replace('\n', '')
	definitions.append(definition)




for i in range(0, len(soup.find_all(class_='definition-table')[0].select('li'))):
	definition =""
	for string in (soup.find_all(class_='definition-table')[0].select('li')[i].strings):
		definition += string
		definition.replace('\n', '')
	definitions.append(definition)

print(definitions)
###########################
## Put definitions into spreadsheet
##
##
##
###########################


